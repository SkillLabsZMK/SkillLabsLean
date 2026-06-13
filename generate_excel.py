"""
Generate Ansprechpartner_Review.xlsx from the SL_DATA embedded in index.html.

The data model uses node trees; we flatten every device to a single row by
extracting the "best" contact / escalation and any steps that appear in the tree.

Strategy per node.kind:
  self    → steps from node.steps; escalation from node.escalation.contact
  contact → no steps; primary contact from node.contact
  branch  → combine both yes/no paths (steps merged; first real contact wins)
  choice  → combine all option paths similarly
  safety  → warn text as a step; contact from node.inner (contact or escalation)
  info    → steps from node.steps; escalation.contact if present
"""

import json, re
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────────────

def _collect(node, steps_acc, contacts_acc, depth=0):
    """Recursively walk a node tree and collect steps + contact dicts."""
    if node is None or depth > 6:
        return
    kind = node.get("kind", "")

    # safety warn → prepend as a step
    if kind == "safety":
        warn = node.get("warn", "")
        if warn:
            steps_acc.append(f"⚠ {warn}")
        _collect(node.get("inner"), steps_acc, contacts_acc, depth + 1)
        return

    # steps from this node
    for s in node.get("steps", []):
        steps_acc.append(s)

    # primary contact
    c = node.get("contact")
    if c and c.get("name"):
        contacts_acc.append({"name": c["name"], "lines": c.get("lines", [])})

    # escalation contact
    esc = node.get("escalation")
    if esc:
        ec = esc.get("contact")
        if ec and ec.get("name"):
            contacts_acc.append({"name": ec["name"], "lines": ec.get("lines", [])})

    # recurse into branch yes/no
    for key in ("yes", "no"):
        child = node.get(key)
        if child:
            _collect(child, steps_acc, contacts_acc, depth + 1)

    # recurse into choice options
    for opt in node.get("options", []):
        _collect(opt.get("node"), steps_acc, contacts_acc, depth + 1)


def flatten_device(room_name, dev):
    """Return a dict with all spreadsheet columns for one device."""
    node = dev.get("node", {})
    steps_acc, contacts_acc = [], []
    _collect(node, steps_acc, contacts_acc)

    # de-duplicate steps while preserving order
    seen = set()
    unique_steps = []
    for s in steps_acc:
        if s not in seen:
            seen.add(s)
            unique_steps.append(s)

    # de-duplicate contacts
    seen_names = set()
    unique_contacts = []
    for c in contacts_acc:
        n = c["name"]
        if n not in seen_names:
            seen_names.add(n)
            unique_contacts.append(c)

    contact1 = unique_contacts[0] if len(unique_contacts) > 0 else {"name": "", "lines": []}
    contact2 = unique_contacts[1] if len(unique_contacts) > 1 else None

    steps_text = " → ".join(f"({i+1}) {s}" for i, s in enumerate(unique_steps)) if unique_steps else ""

    return {
        "raum": room_name,
        "geraet": dev.get("name", ""),
        "status": dev.get("status", ""),
        "selbsthilfe": steps_text,
        "ap1_name": contact1["name"],
        "ap1_detail": " | ".join(contact1["lines"]),
        "ap2_name": contact2["name"] if contact2 else "",
        "ap2_detail": " | ".join(contact2["lines"]) if contact2 else "",
    }


# ── pull all rooms + areas ────────────────────────────────────────────────────

html = open("/home/user/SkillLabsLean/index.html", encoding="utf-8").read()

# Extract ROOMS JSON block
rooms_match = re.search(r"const ROOMS\s*=\s*(\[.*?\]);\s*\n", html, re.DOTALL)
areas_match = re.search(r"const AREAS\s*=\s*(\[.*?\]);\s*\n", html, re.DOTALL)

if not rooms_match:
    raise ValueError("Could not find ROOMS array")
if not areas_match:
    raise ValueError("Could not find AREAS array")

ROOMS = json.loads(rooms_match.group(1))
AREAS = json.loads(areas_match.group(1))

rows = []
for room in ROOMS:
    rname = room["name"]
    for dev in room.get("devices", []):
        rows.append(flatten_device(rname, dev))

for area in AREAS:
    aname = area["name"]
    for dev in area.get("devices", []):
        rows.append(flatten_device(aname, dev))

# ── workbook ─────────────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Ansprechpartner Review"

HEADERS = [
    "Raum",
    "Gerät / Thema",
    "Status",
    "Selbsthilfe-Schritte",
    "1. Ansprechpartner",
    "Kontakt Details",
    "2. Ansprechpartner / Eskalation",
    "Eskalation Details",
    "Korrekt?",
    "Anmerkung",
]

COL_WIDTHS = [20, 30, 12, 45, 28, 35, 25, 35, 12, 30]

# styles
HEADER_FILL  = PatternFill("solid", fgColor="2E4057")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
GRAY_FILL    = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
WRAP         = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

THICK_SIDE   = Side(style="medium", color="2E4057")
THIN_SIDE    = Side(style="thin",   color="CCCCCC")
THICK_TOP    = Border(top=THICK_SIDE)
NORMAL_BORDER= Border(
    top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE
)

# header row
ws.append(HEADERS)
for col_idx, cell in enumerate(ws[1], start=1):
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = HEADER_ALIGN

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# set column widths
for col_idx, width in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# data rows
prev_room = None
for row_num, r in enumerate(rows, start=2):
    room_changed = (r["raum"] != prev_room)
    prev_room = r["raum"]

    ws.append([
        r["raum"],
        r["geraet"],
        r["status"],
        r["selbsthilfe"],
        r["ap1_name"],
        r["ap1_detail"],
        r["ap2_name"],
        r["ap2_detail"],
        "",   # Korrekt?
        "",   # Anmerkung
    ])

    is_even = (row_num % 2 == 0)
    fill = GRAY_FILL if is_even else WHITE_FILL

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill      = fill
        cell.alignment = WRAP
        # thicker top border when room changes
        if room_changed:
            cell.border = Border(
                top=THICK_SIDE,
                bottom=THIN_SIDE,
                left=THIN_SIDE,
                right=THIN_SIDE,
            )
        else:
            cell.border = NORMAL_BORDER

# row heights – auto-size is not supported in openpyxl; set a sensible default
ws.row_dimensions[1].height = 30
for row_num in range(2, len(rows) + 2):
    ws.row_dimensions[row_num].height = 60

# save
out_path = "/home/user/SkillLabsLean/Ansprechpartner_Review.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Data rows written: {len(rows)}")
