"""
Apply reviewed Excel → index.html.
- Reads AP1 (col E/F) and AP2 (col G/H)
- Stores ap2 in node for dual-box rendering
- Handles device removals and renames
"""

import json, re
from openpyxl import load_workbook

EXCEL_PATH = "/root/.claude/uploads/1a95ad9a-4c85-5103-8d4b-289795273960/a49d8529-Ansprechpartner_Review.xlsx"
HTML_PATH  = "/home/user/SkillLabsLean/index.html"

# Devices to explicitly delete (reviewer removed from Excel)
TO_DELETE = [
    ("Seminar 1",         "Medienpanel (Extron)"),
    ("Seminar 3",         "Kamera / Medientechnik defekt"),
    ("Technicum",         "Demokamera"),
    ("Technicum",         "Notschalter ausgelöst"),
    ("Gipsen & Polieren", "Erkoform 3d motion"),   # renamed → new entry adds updated name
]

# ── Read Excel ────────────────────────────────────────────────────────────────

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb.active

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    raum     = (row[0] or "").strip()
    geraet   = (row[1] or "").strip()
    if not raum or not geraet:
        continue
    rows.append({
        "raum":     raum,
        "geraet":   geraet,
        "status":   (row[2] or "").strip(),
        "schritte": (row[3] or "").strip(),
        "ap1_name": (row[4] or "").strip(),
        "ap1_det":  (row[5] or "").strip(),
        "ap2_name": (row[6] or "").strip(),
        "ap2_det":  (row[7] or "").strip(),
    })

print(f"Excel rows: {len(rows)}")

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_steps(text):
    if not text:
        return []
    parts = re.split(r" → (?=\(\d+\) )", text)
    return [re.sub(r"^\(\d+\)\s*", "", p.strip()) for p in parts if p.strip()]

def split_lines(detail):
    return [l.strip() for l in detail.split("|") if l.strip()] if detail else []

def build_node(device_name, ap1_name, ap1_det, ap2_name, ap2_det, schritte_text):
    steps = parse_steps(schritte_text)
    lines  = split_lines(ap1_det)
    lines2 = split_lines(ap2_det)

    # Safety warn: first step starts with ⚠
    safety_warn  = None
    inner_steps  = []
    for i, step in enumerate(steps):
        if i == 0 and step.startswith("⚠"):
            safety_warn = step.lstrip("⚠").strip()
        else:
            inner_steps.append(step)

    # Build contact objects
    ap2 = {"name": ap2_name, "lines": lines2} if ap2_name else None

    if ap1_name and inner_steps:
        esc_node = {"kind": "contact", "contact": {"name": ap1_name, "lines": lines}, "desc": "", "pills": []}
        if ap2:
            esc_node["ap2"] = ap2
        core = {"kind": "self", "title": device_name, "steps": inner_steps, "escalation": esc_node}
    elif ap1_name:
        core = {"kind": "contact", "contact": {"name": ap1_name, "lines": lines}, "desc": "", "pills": []}
        if ap2:
            core["ap2"] = ap2
    elif inner_steps:
        core = {"kind": "self", "title": device_name, "steps": inner_steps}
    else:
        core = {"kind": "contact", "contact": {"name": "", "lines": lines}}

    if safety_warn:
        return {"kind": "safety", "warn": safety_warn, "inner": core}
    return core

# ── Read HTML ─────────────────────────────────────────────────────────────────

html = open(HTML_PATH, encoding="utf-8").read()

rooms_m = re.search(r"(const ROOMS\s*=\s*)(\[.*?\])(\s*;\s*\n)", html, re.DOTALL)
areas_m = re.search(r"(const AREAS\s*=\s*)(\[.*?\])(\s*;\s*\n)", html, re.DOTALL)
if not rooms_m or not areas_m:
    raise ValueError("Cannot find ROOMS or AREAS")

ROOMS = json.loads(rooms_m.group(2))
AREAS = json.loads(areas_m.group(2))
ALL   = ROOMS + AREAS

def find_device(raum, geraet):
    for r in ALL:
        if r["name"] == raum:
            for d in r.get("devices", []):
                if d["name"] == geraet:
                    return r, d
    return None, None

# ── Delete removed devices ────────────────────────────────────────────────────

for raum, geraet in TO_DELETE:
    for r in ALL:
        if r["name"] == raum:
            before = len(r["devices"])
            r["devices"] = [d for d in r["devices"] if d["name"] != geraet]
            if len(r["devices"]) < before:
                print(f"Deleted: [{raum}] {geraet}")

# ── Apply updates ─────────────────────────────────────────────────────────────

updated, not_found = 0, []

for row in rows:
    _, dev = find_device(row["raum"], row["geraet"])
    if dev is None:
        not_found.append(row)
        continue
    dev["node"] = build_node(
        row["geraet"], row["ap1_name"], row["ap1_det"],
        row["ap2_name"], row["ap2_det"], row["schritte"]
    )
    if row["status"]:
        dev["status"] = row["status"]
    updated += 1

print(f"Updated: {updated}  |  Not found: {len(not_found)}")
for r in not_found:
    print(f"  [{r['raum']}] {r['geraet']}")

for row in not_found:
    for r in ALL:
        if r["name"] == row["raum"]:
            node = build_node(
                row["geraet"], row["ap1_name"], row["ap1_det"],
                row["ap2_name"], row["ap2_det"], row["schritte"]
            )
            r["devices"].append({
                "name":    row["geraet"],
                "status":  row["status"] or node["kind"],
                "node":    node,
                "manuals": []
            })
            print(f"  Added: [{row['raum']}] {row['geraet']}")
            break

# ── Write back ────────────────────────────────────────────────────────────────

rooms_json = json.dumps(ROOMS, ensure_ascii=False, indent=2)
areas_json = json.dumps(AREAS, ensure_ascii=False, indent=2)

rooms_new  = rooms_m.group(1) + rooms_json + rooms_m.group(3)
html       = html[:rooms_m.start()] + rooms_new + html[rooms_m.end():]

areas_m2   = re.search(r"(const AREAS\s*=\s*)(\[.*?\])(\s*;\s*\n)", html, re.DOTALL)
areas_new  = areas_m2.group(1) + areas_json + areas_m2.group(3)
html       = html[:areas_m2.start()] + areas_new + html[areas_m2.end():]

with open(HTML_PATH, "w", encoding="utf-8") as f:
    f.write(html)

print("Done.")
